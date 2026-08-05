"""Chargement et scoring du catalogue de datasets (fichier metadonnees_datasets.xlsx) :
résolution des identifiants, tokens de recherche, formats d'export."""
from __future__ import annotations

import json
import re
import time
import unicodedata
import io
import tempfile
import zipfile
import sqlite3
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import requests
from flask import Flask, jsonify, render_template, request, send_from_directory, send_file

from connectors import DataGouvConnector, DataGouvConnectorError
from history import scheduler as history_scheduler


# BASE_DIR recalculé : ce fichier vit dans scripts/, la racine du projet
# (contenant data/, templates/, static/...) est donc le dossier PARENT de scripts/.
BASE_DIR = Path(__file__).resolve().parent.parent

from .open_data_api import *  # noqa: F401,F403
from .local_sources import *  # noqa: F401,F403


CATALOG_XLSX_PATH = BASE_DIR / "data" / "metadonnees_datasets.xlsx"
CATALOG_CACHE: Optional[list[Dict[str, Any]]] = None

def _normalize_title_for_match(value: Any) -> str:
    """Normalise un titre afin de fiabiliser les comparaisons entre jeux de données."""
    return re.sub(r"[^a-z0-9]+", " ", _strip_accents(value or "").lower()).strip()


def _local_dataset_id_for_title(title: str) -> Optional[str]:
    """Match a catalogue row title against LOCAL_DATASETS when its URL isn't a
    recognized OpenDataSoft explore link (ex. fichiers locaux type INSEE)."""
    normalized_title = _normalize_title_for_match(title)
    if not normalized_title:
        return None
    for local_id, config in LOCAL_DATASETS.items():
        if _normalize_title_for_match(config.get("title", "")) == normalized_title:
            return local_id
    return None


def _slug_from_catalog_url(value: Any) -> Optional[str]:
    """Extrait le slug d'un jeu de données à partir de son URL de catalogue."""
    text = str(value or "")
    match = re.search(r"https?://([^/]+)/explore/dataset/([^/?#]+)", text, re.I)
    if not match:
        match = re.search(r"https?://([^/]+)/(?:api/explore/v2(?:\.1)?/)?catalog/datasets/([^/?#]+)", text, re.I)
    if not match:
        return None
    domain = match.group(1).strip().lower()
    slug = match.group(2).strip().lower()
    if not DATASET_ID_PATTERN.fullmatch(slug) or not ODS_DOMAIN_PATTERN.fullmatch(domain):
        return None
    return _make_ods_dataset_id(domain, slug)


def _catalog_tokens(*values: Any) -> set[str]:
    """Construit l'ensemble des mots normalisés utilisés pour comparer les entrées du catalogue."""
    text = _strip_accents(" ".join(str(v or "") for v in values)).lower()
    tokens = set(re.findall(r"[a-z0-9]{2,}", text))
    stop = {"de", "des", "du", "la", "le", "les", "en", "et", "pour", "sur", "avec", "dans", "data", "donnees", "dataset", "toulouse", "metropole", "open"}
    return tokens - stop


def _load_catalog_entries() -> list[Dict[str, Any]]:
    """Charge et normalise les entrées du catalogue de métadonnées."""
    global CATALOG_CACHE
    if CATALOG_CACHE is not None:
        return CATALOG_CACHE
    if not CATALOG_XLSX_PATH.exists():
        CATALOG_CACHE = []
        return CATALOG_CACHE

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(CATALOG_XLSX_PATH, read_only=False, data_only=False)
    except Exception:
        CATALOG_CACHE = []
        return CATALOG_CACHE

    entries: list[Dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower().startswith("présentation") or sheet_name.lower().startswith("presentation"):
            continue
        sheet = workbook[sheet_name]
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        normalized_headers = [_normalize_field_name(h) for h in headers]
        url_indexes = [i for i, h in enumerate(normalized_headers) if h == "url"]
        geo_indexes = [i for i, h in enumerate(normalized_headers) if "geolocalisation" in h or h in {"geo", "geometry", "geometrie"}]
        attribute_indexes = [i for i, h in enumerate(normalized_headers) if "attribut" in h]
        comment_indexes = [i for i, h in enumerate(normalized_headers) if "commentaire" in h or "description" in h]
        tag_indexes = [i for i, h in enumerate(normalized_headers) if h.startswith("tag")]
        format_indexes = [i for i, h in enumerate(normalized_headers) if h == "format"]

        for row in sheet.iter_rows(min_row=2):
            title = str(row[0].value or "").strip()
            if not title:
                continue
            url = ""
            for idx in url_indexes:
                cell = row[idx]
                if cell.hyperlink and cell.hyperlink.target:
                    url = cell.hyperlink.target
                    break
                if cell.value:
                    url = str(cell.value)
                    break
            dataset_id = _slug_from_catalog_url(url) or _local_dataset_id_for_title(title)
            if not dataset_id:
                continue
            geo_text = " ".join(str(row[i].value or "") for i in geo_indexes)
            attrs = " ".join(str(row[i].value or "") for i in attribute_indexes)
            comments = " ".join(str(row[i].value or "") for i in comment_indexes)
            tags = " ".join(str(row[i].value or "") for i in tag_indexes)
            raw_format = " ".join(str(row[i].value or "") for i in format_indexes).strip()
            negative_geo = bool(re.search(r"\b(pas|sans|aucun(?:e)?)\s+(?:de\s+)?(?:geo|geometr|coordonn|localis)", _strip_accents(geo_text + " " + comments).lower()))
            geo_hint = bool(re.search(r"geopoint|geoshape|latitude|longitude|geometry|geometrie|coordonne", _strip_accents(geo_text + " " + attrs).lower())) and not negative_geo
            entries.append({
                "dataset_id": dataset_id,
                "title": title,
                "theme": sheet_name,
                "url": url,
                "geo_hint": geo_hint,
                "geo_text": geo_text,
                "attributes": attrs,
                "comments": comments,
                "tags": tags,
                "format": raw_format,
                "tokens": _catalog_tokens(title, sheet_name, comments, tags, attrs),
            })
    CATALOG_CACHE = entries
    return entries


def _scalar_values(records: list[Dict[str, Any]], field: str, limit: int = 500) -> list[Any]:
    """Return the non-empty values of `field` across records, capped at `limit`."""
    values: list[Any] = []
    for record in records:
        if field not in record:
            continue
        value = record.get(field)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        values.append(value)
        if len(values) >= limit:
            break
    return values


def _slugify_for_filename(text: str) -> str:
    """Transforme un texte en nom de fichier sûr et lisible."""
    normalized = _strip_accents(text or "dataset").lower()
    slug = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return slug or "dataset"


def _catalog_format_for_dataset(dataset_id: str) -> str:
    """Return the raw 'Format' value declared in the catalogue for a dataset."""
    for entry in _load_catalog_entries():
        if entry["dataset_id"] == dataset_id:
            return entry.get("format", "") or ""
    return ""


EXPORT_DRIVERS = {
    "gpkg": "GPKG",
    "shp": "ESRI Shapefile",
    "kml": "KML",
}


def _normalize_export_format(raw_format: str) -> str:
    """Map a catalogue 'Format' string to one of our supported export kinds.

    Falls back to GeoJSON — a lossless, universally supported geographic
    format — whenever the declared format is ambiguous, mixed, or unknown.
    """
    text = _strip_accents(raw_format or "").lower()
    if "gpkg" in text or "geopackage" in text:
        return "gpkg"
    if "shapefile" in text or re.search(r"\bshp\b", text):
        return "shp"
    if re.search(r"\bkml\b", text):
        return "kml"
    if "geojson" in text or "json" in text or "/" in text:
        # Multiple formats declared together, or already geo-native: keep the
        # lossless geographic format rather than guessing which one "wins".
        return "geojson"
    if "parquet" in text:
        return "parquet"
    if "xlsx" in text or "excel" in text:
        return "xlsx"
    if re.search(r"\bcsv\b", text):
        return "csv"
    return "geojson"



    values = []
    for record in records[:limit]:
        value = record.get(field)
        if isinstance(value, (dict, list, tuple, set)) or value in (None, ""):
            continue
        values.append(value)
    return values




__all__ = ['CATALOG_CACHE', 'CATALOG_XLSX_PATH', 'EXPORT_DRIVERS', '_catalog_format_for_dataset', '_catalog_tokens', '_load_catalog_entries', '_local_dataset_id_for_title', '_normalize_export_format', '_normalize_title_for_match', '_scalar_values', '_slug_from_catalog_url', '_slugify_for_filename']
